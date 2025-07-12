from pywinauto import Application
import time
from methods_to_df import  exportar_reporte_HK,exportar_reporte_DFT,exportar_reporte_BJH_con_teclas,exportar_reporte_fractal_con_teclas,exportar_reporte_BET_con_teclas
import os
import pandas as pd
import configparser
import openpyxl
from pywinauto import Application
import traceback
import subprocess
from pywinauto.keyboard import send_keys
from datetime import datetime
from queue import Queue
import threading

def inicializar_novawin(path_novawin):
    for backend in ["win32", "uia"]:
        try:
            print(f"Intentando iniciar NovaWin con backend='{backend}'...")
            app = Application(backend=backend).start(path_novawin)
            time.sleep(10)  # espera para que cargue

            main_window = app.window(title_re=".*NovaWin.*")
            main_window.wait("exists enabled visible ready", timeout=15)

            # Intentar mover ventana fuera de pantalla en lugar de minimizarla
            try:
                # Cambiar tamaño y posición de la ventana (x=0, y=0, width=800, height=600)
                main_window.move_window(x=0, y=0, width=800, height=600, repaint=True)
                print("Ventana redimensionada y movida")
   
            except AttributeError:
                print("Backend no soporta move_window")

            # NO llamar a set_focus() si está minimizada (puede bloquear)
            if main_window.is_minimized():
                print("La ventana está minimizada, la dejaremos así para no bloquear")
            else:
                main_window.set_focus()

            print(f"Ventana NovaWin lista con backend='{backend}'")
            return app, main_window

        except Exception as e:
            print(f"Error con backend '{backend}': {e}")

    raise RuntimeError("No se pudo iniciar NovaWin con ninguno de los backends disponibles.")
def crear_excel_con_hojas(csv_dict, ruta_excel_final):
    with pd.ExcelWriter(ruta_excel_final, engine='openpyxl') as writer:
        for hoja, path_csv in csv_dict.items():
            try:
                df = pd.read_csv(path_csv)
                df.to_excel(writer, sheet_name=hoja, index=False)
                print(f" Hoja '{hoja}' agregada con éxito desde {path_csv}")
            except Exception as e:
                print(f" Error al procesar {hoja} ({path_csv}): {e}")


def manejar_novawin(path_qps, path_csv, path_novawin, carpeta_destino):
    import os
    import pandas as pd
    import time
    import configparser
    from datetime import datetime
    from pywinauto.keyboard import send_keys

    try:
        path_qps = os.path.normpath(path_qps)
        path_csv = os.path.normpath(path_csv)
        carpeta_destino = os.path.abspath(carpeta_destino)

        if not os.path.isdir(carpeta_destino):
            raise Exception(f"La ruta no es una carpeta válida: {carpeta_destino}")

        # Generar nombre único para el archivo planilla
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"informe_completo_{timestamp}.xlsx"
        archivo_planilla = os.path.join(carpeta_destino, nombre_archivo)

        # Guardar en config.ini
        config_file = "config.ini"
        config = configparser.ConfigParser()
        config.read(config_file)
        if "Rutas" not in config:
            config["Rutas"] = {}
        config["Rutas"]["archivo_planilla"] = archivo_planilla
        with open(config_file, "w") as f:
            config.write(f)

        print("Iniciando NovaWin...")
        app, main_window = inicializar_novawin(path_novawin)
        time.sleep(2)

        print("Abriendo archivo .QPS con Alt+F, luego O...")
        send_keys('%fo')
        time.sleep(1.5)

        print(f"Ingresando ruta del archivo QPS: {path_qps}")
        send_keys(path_qps)
        time.sleep(0.5)
        send_keys('%a')
        print("Archivo .QPS enviado y abierto.")
        time.sleep(2)

        # Exportar todos los reportes
        path_csv_hk = exportar_reporte_HK(main_window, carpeta_destino, app)
        path_csv_dft = exportar_reporte_DFT(main_window, carpeta_destino, app)
        path_csv_bjhad = exportar_reporte_BJH_con_teclas(main_window, carpeta_destino, app)
        path_csv_n = exportar_reporte_fractal_con_teclas(main_window, carpeta_destino, app, "n")
        path_csv_f = exportar_reporte_fractal_con_teclas(main_window, carpeta_destino, app, "f")
        path_csv_k = exportar_reporte_fractal_con_teclas(main_window, carpeta_destino, app, "k")
        path_csv_h = exportar_reporte_fractal_con_teclas(main_window, carpeta_destino, app, "h")
        path_csv_bet = exportar_reporte_BET_con_teclas(main_window, carpeta_destino, app)

        if app.is_process_running():
         app.kill()
         print("NovaWin cerrado correctamente.")
         
        import pandas as pd
        import os

        # Diccionario con nombres de hoja y rutas de los CSV
        csv_hojas = {
        "HK": path_csv_hk,
        "DFT": path_csv_dft,
        "BJHA": path_csv_bjhad[0],
        "BJHD": path_csv_bjhad[1],
        "N": path_csv_n,
        "F": path_csv_f,
        "K": path_csv_k,
        "H": path_csv_h,
        "BET": path_csv_bet,
        }

        # Ruta del Excel final
        ruta_final = os.path.join(os.path.dirname(archivo_planilla), "informe_completo.xlsx")

        # Crear Excel con varias hojas
        with pd.ExcelWriter(ruta_final, engine='openpyxl') as writer:
          for nombre_hoja, path_csv in csv_hojas.items():
             if not os.path.isfile(path_csv):
                print(f"Ignorado: {path_csv} no es un archivo válido.")
                continue
             try:
               df = pd.read_csv(path_csv)
               df.to_excel(writer, sheet_name=nombre_hoja, index=False)
             except Exception as e:
               print(f"No se pudo leer {path_csv}: {e}")

          print(f"Informe con hojas múltiples guardado en: {ruta_final}")
          config["Rutas"]["archivo_informe"] = ruta_final
          with open(config_file, "w") as f:
              config.write(f)
              # Al final de manejar_novawin:
          try:
            import tkinter as tk
            ventana = tk._default_root  # Recupera la ventana si aún está activa
            if ventana:
                ventana.geometry("600x600")
                ventana.update_idletasks()
                ventana.lift()
                ventana.focus_force()
        except Exception as e:
            print(f"No se pudo restablecer la ventana principal: {e}")
    
    except Exception as e:
        print(f"Error al manejar NovaWin: {e}")
def seleccionar_menu(window, ruta_menu):
    try:
        print(f"Seleccionando menú: {ruta_menu}")
        window.menu_select(ruta_menu)
        time.sleep(2)
    except Exception as e:
        print(f"Error al seleccionar menú '{ruta_menu}': {e}")
        raise

def interactuar_con_cuadro_dialogo(dialog, archivo):
    try:
        print(f"Interactuando con diálogo para abrir: {archivo}")
        edit_box = dialog.child_window(class_name="Edit")
        edit_box.set_edit_text(archivo)
        
        open_button = dialog.child_window(class_name="Button", found_index=0)
        open_button.click_input()
        
        print("Archivo enviado al diálogo correctamente.")
    except Exception as e:
        print(f"Error al interactuar con el cuadro de diálogo: {e}")
        raise

def leer_csv_y_crear_dataframe(ruta_csv):
    if not os.path.exists(ruta_csv):
        print(f"Archivo CSV no encontrado: {ruta_csv}")
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_csv}")
    try:
        df = pd.read_csv(ruta_csv)
        print(f"CSV leído correctamente: {ruta_csv}")
        return df
    except Exception as e:
        print(f"Error al leer CSV: {e}")
        raise

def agregar_csv_a_plantilla_excel(ruta_csv, ruta_excel, df_csv):
    try:
        ruta_excel = os.path.normpath(os.path.join(ruta_excel, "Report.xlsx"))
        print(f"Archivo Excel destino: {ruta_excel}")

        if not os.path.exists(ruta_excel) or not ruta_excel.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            wb.save(ruta_excel)
            print(f"Archivo Excel creado: {ruta_excel}")

        wb = openpyxl.load_workbook(ruta_excel)
        ws = wb["Reporte"] if "Reporte" in wb.sheetnames else wb.create_sheet("Reporte")

        max_row = ws.max_row
        max_col = ws.max_column
        start_col = max_col + 1 if max_row > 1 else 1

        if start_col == 1:
            for col, header in enumerate(df_csv.columns, start=start_col):
                ws.cell(row=1, column=col).value = header

        for i, row in enumerate(df_csv.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=start_col):
                ws.cell(row=i, column=j).value = value

        wb.save(ruta_excel)
        print(f"Datos del CSV agregados a: {ruta_excel}")
    except Exception as e:
        print(f"Error al agregar datos a Excel: {e}")
        raise

def guardar_dataframe_en_ini(df, archivo_ini):
    try:
        config = configparser.ConfigParser()
        for columna in df.columns:
            config[columna] = {f"fila_{i}": str(valor) for i, valor in enumerate(df[columna])}
        with open(archivo_ini, 'w') as archivo:
            config.write(archivo)
        print(f"DataFrame guardado en INI: {archivo_ini}")
    except Exception as e:
        print(f"Error al guardar INI: {e}")
        raise

def close_window_novawin():
    try:
        app = Application(backend='uia').connect(title_re='.*NovaWin.*')
        window = app.window(title_re='.*NovaWin.*')
        window.close()
        print("La ventana de NovaWin ha sido cerrada.")
    except Exception as e:
        print(f"Error al cerrar la ventana de NovaWin: {e}")

def ejecutar_ide():
    try:
        subprocess.run(["python", "-m", "novarep_ide"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error ejecutando IDE: {e}")
