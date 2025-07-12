from tkinter import Tk, Label, Button, Entry, filedialog, Frame, Scrollbar, VERTICAL, HORIZONTAL, RIGHT, Y, BOTTOM, X, BOTH  
from tkinter import ttk
import configparser
import pandas as pd
from tests import tests_main
from config_ini import main as config_main
import subprocess
from graphs import draw_HK,draw_DFT,draw_comparison_bar_chart  # Asegúrate de importar tu función
from tkinter import Toplevel
from PIL import Image, ImageTk
import os
from cesarofrac import visualizar_poro_fractal
config_file = "config.ini"

ventana = Tk()
ventana.title("Selector de Archivo Excel")
ventana.geometry("1000x1100")
ventana.resizable(False, False)
# --- Estado ---
label_estado = Label(ventana, text="", fg="blue", font=("Arial", 10))
label_estado.grid(row=2, column=1, columnspan=2, pady=5, sticky="w")
# --- Ruta archivo Excel ---
Label(ventana, text="Ruta del archivo Excel:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
entry_excel = Entry(ventana, width=50)
entry_excel.grid(row=0, column=1, padx=10, pady=10)
ruta_img_dft=""
ruta_img_hk=""
ruta_img_bjhd=""
ruta_img_bjha=""
from tkinter import Toplevel, Label, Canvas, Scrollbar, Frame, VERTICAL, RIGHT, LEFT, BOTH, Y
from PIL import Image, ImageTk
import os
import configparser

def mostrar_imagenes_en_carrusel():
    try:
        config = configparser.ConfigParser()
        config.read("startconfig.ini")
        if "Rutas" in config and "ruta_excel" in config["Rutas"]:
            directorio = config["Rutas"]["ruta_excel"]
            if not os.path.isdir(directorio):
                label_estado.config(text="Directorio inválido.")
                return

            imagenes = [f for f in os.listdir(directorio) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
            imagenes = sorted(imagenes)

            if not imagenes:
                label_estado.config(text="No se encontraron imágenes.")
                return

            ventana_carousel = Toplevel()
            ventana_carousel.title("Carrusel de Imágenes")
            ventana_carousel.geometry("700x500")

            indice_actual = [0]  # lista mutable para poder modificar dentro de funciones internas
            imagenes_tk = []

            # Cargar y guardar todas las imágenes redimensionadas
            for nombre in imagenes:
                img_path = os.path.join(directorio, nombre)
                img = Image.open(img_path)
                img = img.resize((600, 340), Image.Resampling.LANCZOS)
                img_tk = ImageTk.PhotoImage(img)
                imagenes_tk.append((nombre, img_tk))

            # Widgets
            label_nombre = Label(ventana_carousel, text=imagenes_tk[0][0], font=("Arial", 12, "bold"))
            label_nombre.pack(pady=10)

            label_img = Label(ventana_carousel, image=imagenes_tk[0][1])
            label_img.pack()

            def mostrar_imagen(indice):
                nombre, img_tk = imagenes_tk[indice]
                label_img.configure(image=img_tk)
                label_img.image = img_tk  # evitar garbage collection
                label_nombre.config(text=nombre)

            def imagen_anterior():
                if indice_actual[0] > 0:
                    indice_actual[0] -= 1
                    mostrar_imagen(indice_actual[0])

            def imagen_siguiente():
                if indice_actual[0] < len(imagenes_tk) - 1:
                    indice_actual[0] += 1
                    mostrar_imagen(indice_actual[0])

            # Botones
            btn_izquierda = Button(ventana_carousel, text="Anterior", command=imagen_anterior)
            btn_izquierda.pack(side=LEFT, padx=20, pady=20)

            btn_derecha = Button(ventana_carousel, text="Siguiente️", command=imagen_siguiente)
            btn_derecha.pack(side=RIGHT, padx=20, pady=20)

        else:
            label_estado.config(text="No se encontró 'ruta_excel' en startconfig.ini")
    except Exception as e:
        label_estado.config(text=f"Error: {e}")
def hacer_tests(ruta_excel, label_estado):
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image

    if not os.path.isfile(ruta_excel):
        label_estado.config(text="El archivo no existe.")
        return

    try:
        # Leer hoja BET como DataFrame
        df_bet = pd.read_excel(ruta_excel, sheet_name="BET")

        # Ejecutar análisis (devuelve df de resultados y ruta del gráfico)
        resultados_df, ruta_grafico = tests_main(df_bet, os.path.dirname(ruta_excel))

        # Abrir Excel y acceder a la hoja "BET"
        wb = load_workbook(ruta_excel)
        if "BET" not in wb.sheetnames:
            wb.create_sheet("BET")
        hoja_bet = wb["BET"]

        # Buscar la primera fila vacía
        fila_inicial = hoja_bet.max_row + 2

        # Agregar resultados desde la fila vacía
        for r_idx, row in enumerate(dataframe_to_rows(resultados_df, index=False, header=True)):
            for c_idx, value in enumerate(row, 1):
                hoja_bet.cell(row=fila_inicial + r_idx, column=c_idx, value=value)

        # Insertar gráfico debajo de los resultados
        if os.path.exists(ruta_grafico):
            img = Image(ruta_grafico)
            img.anchor = f"G{fila_inicial + len(resultados_df) + 2}"
            hoja_bet.add_image(img)

        # Guardar cambios
        wb.save(ruta_excel)
        wb.close()

        # Eliminar imagen temporal
        if os.path.exists(ruta_grafico):
            os.remove(ruta_grafico)

        label_estado.config(text="Resultados de BET insertados correctamente.")

    except Exception as e:
        label_estado.config(text=f"Error al procesar BET: {e}")
def hacer_BJHS(ruta_excel, label_estado):
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

    if not os.path.isfile(ruta_excel):
        label_estado.config(text="El archivo no existe.")
        return

    try:
        df_bjha = pd.read_excel(ruta_excel, sheet_name="BJHA")
        df_bjhd = pd.read_excel(ruta_excel, sheet_name="BJHD")

        # Obtener rutas de imágenes separadas
        ruta_img_bjhd, ruta_img_bjha = draw_comparison_bar_chart(df_bjhd, df_bjha)

        wb = load_workbook(ruta_excel)

        # --- Insertar en hoja BJHD ---
        if "BJHD" not in wb.sheetnames:
            wb.create_sheet("BJHD")
        hoja_bjhd = wb["BJHD"]
        fila_destino_bjhd = hoja_bjhd.max_row + 2

        img_bjhd = Image(ruta_img_bjhd)
        img_bjhd.anchor = f"A{fila_destino_bjhd}"
        hoja_bjhd.add_image(img_bjhd)

        # --- Insertar en hoja BJHA ---
        if "BJHA" not in wb.sheetnames:
            wb.create_sheet("BJHA")
        hoja_bjha = wb["BJHA"]
        fila_destino_bjha = hoja_bjha.max_row + 2

        img_bjha = Image(ruta_img_bjha)
        img_bjha.anchor = f"A{fila_destino_bjha}"
        hoja_bjha.add_image(img_bjha)

        # Guardar cambios
        wb.save(ruta_excel)
        wb.close()
        # Al final de hacer_BJHS():
        if os.path.exists("reports") and os.path.isdir("reports"):
            shutil.move(ruta_img_bjhd, os.path.join("reports", os.path.basename(ruta_img_bjhd)))
            shutil.move(ruta_img_bjha, os.path.join("reports", os.path.basename(ruta_img_bjha)))
        # Eliminar imágenes temporales
        #os.remove(ruta_img_bjhd)
        #os.remove(ruta_img_bjha)

        label_estado.config(text="Gráficos BJHD y BJHA generados en sus respectivas hojas.")

    except Exception as e:
        label_estado.config(text=f"Error: {e}")
def hacer_DFT(ruta_excel, label_estado):
    import os
    import shutil
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

    if not os.path.isfile(ruta_excel):
        label_estado.config(text="El archivo no existe.")
        return

    try:
        df_dft = pd.read_excel(ruta_excel, sheet_name="DFT")
        ruta_imagen_dft = draw_DFT(df_dft)

        wb = load_workbook(ruta_excel)
        if "DFT" not in wb.sheetnames:
            wb.create_sheet("DFT")
        hoja = wb["DFT"]

        fila_destino = hoja.max_row + 2
        celda_destino = f"A{fila_destino}"

        img = Image(ruta_imagen_dft)
        img.anchor = celda_destino
        hoja.add_image(img)

        wb.save(ruta_excel)
        wb.close()

        # Mover imagen a carpeta "reports"
        if os.path.exists("reports") and os.path.isdir("reports"):
            nombre = os.path.basename(ruta_imagen_dft)
            shutil.move(ruta_imagen_dft, os.path.join("reports", nombre))

        label_estado.config(text="Gráfico DFT generado e insertado.")
    except Exception as e:
        label_estado.config(text=f"Error: {e}")
def hacer_HK(ruta_excel, label_estado):
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

    if not os.path.isfile(ruta_excel):
        label_estado.config(text="El archivo no existe.")
        return

    try:
        # Leer hoja HK
        df_hk = pd.read_excel(ruta_excel, sheet_name="HK")

        # Generar gráfico y obtener ruta
        ruta_imagen_hk = draw_HK(df_hk)

        # Insertar imagen en la hoja "HK"
        wb = load_workbook(ruta_excel)
        if "HK" not in wb.sheetnames:
            wb.create_sheet("HK")
        hoja = wb["HK"]

        # Calcular fila vacía para no sobrescribir datos
        fila_destino = hoja.max_row + 2
        celda_destino = f"A{fila_destino}"

        img = Image(ruta_imagen_hk)
        img.anchor = celda_destino
        hoja.add_image(img)

        wb.save(ruta_excel)
        wb.close()

        # Eliminar imagen temporal
        #os.remove(ruta_imagen_hk)

        label_estado.config(text="Gráfico HK generado e insertado.")
    except Exception as e:
        label_estado.config(text=f"Error: {e}")
def abrir_config_ini():

    try:
     config_main()
     print("Módulo configuración abierto correctamente")
     ventana.destroy()
    except Exception as e:
     print(f"Error al abrir módulo: {e}")
     try:
        label_estado.config(text=f"Error: {e}")
     except:
        pass
def seleccionar_archivo():

    ventana.geometry("1200x1200")        # Cambia el tamaño
    ventana.update_idletasks()          # Fuerza la actualización visual
    ventana.resizable(False, False)     # Opcional: evitar que el usuario redimensione
    archivo = filedialog.askopenfilename(filetypes=[("Archivos XLSX", "*.xlsx")])
    if archivo:
        
        entry_excel.delete(0, 'end')
        entry_excel.insert(0, archivo)

Button(ventana, text="Seleccionar archivo", command=seleccionar_archivo).grid(row=0, column=2, padx=10, pady=10)

# --- Selección de hoja ---
Label(ventana, text="Selecciona hoja:").grid(row=3, column=0, padx=10, pady=10, sticky="w")
combo_hojas = ttk.Combobox(ventana, state="readonly")
combo_hojas.grid(row=4, column=1, padx=10, pady=10, sticky="w")
combo_hojas.grid_forget()

# --- Frame para tabla ---
frame_tabla = Frame(ventana, width=960, height=600)
frame_tabla.grid(row=4, column=0, columnspan=3, padx=10, pady=10)
frame_tabla.grid_forget()

scrollbar_y = Scrollbar(frame_tabla, orient=VERTICAL)
scrollbar_y.pack(side=RIGHT, fill=Y)
scrollbar_x = Scrollbar(frame_tabla, orient=HORIZONTAL)
scrollbar_x.pack(side=BOTTOM, fill=X)

tree = ttk.Treeview(frame_tabla, yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
tree.pack(expand=True, fill=BOTH)

scrollbar_y.config(command=tree.yview)
scrollbar_x.config(command=tree.xview)

def mostrar_dataframe(df):
    combo_hojas.grid(row=3, column=1, padx=10, pady=10, sticky="w")
    frame_tabla.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

    tree.delete(*tree.get_children())
    tree["columns"] = list(df.columns)
    tree["show"] = "headings"

    ancho_fijo = 1000
    column_width = ancho_fijo // len(df.columns)

    for col in df.columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=column_width)

    for _, row in df.iterrows():
        tree.insert("", "end", values=list(row))

def cargar_archivo():
    ruta_excel = entry_excel.get()
    if not ruta_excel:
        label_estado.config(text="Por favor, proporciona la ruta del archivo Excel")
        return
    try:
        hojas = pd.ExcelFile(ruta_excel).sheet_names
        combo_hojas["values"] = hojas
        if hojas:
            combo_hojas.current(0)
            cargar_hoja(ruta_excel, hojas[0])
        label_estado.config(text="Archivo cargado correctamente")
    except Exception as e:
        label_estado.config(text=f"Error: {e}")

def cargar_hoja(ruta_excel, hoja):
    try:
        df = pd.read_excel(ruta_excel, sheet_name=hoja)
        mostrar_dataframe(df)
        label_estado.config(text=f"Mostrando hoja: {hoja}")
    except Exception as e:
        label_estado.config(text=f"Error al cargar hoja: {e}")

def on_hoja_seleccionada(event):
    ruta_excel = entry_excel.get()
    hoja = combo_hojas.get()
    if ruta_excel and hoja:
        cargar_hoja(ruta_excel, hoja)

combo_hojas.bind("<<ComboboxSelected>>", on_hoja_seleccionada)

# Botón para cargar Excel y mostrar hojas
Button(ventana, text="Visualizar Excel", command=cargar_archivo).grid(row=5, column=0, columnspan=3, pady=10)

Button(ventana, text="Ir a Configuración", command=lambda: abrir_config_ini()).grid(row=6, column=0, pady=20)
# Botón hacer HK
Button(
    ventana,
    text="hacer HK",
    command=lambda: hacer_HK(entry_excel.get(), label_estado)
).grid(row=6, column=1, pady=10)

# Botón hacer DFT

Button(
    ventana,
    text="hacer DFT",
    command=lambda: hacer_DFT(entry_excel.get(), label_estado)
).grid(row=6, column=2, pady=10)

# Botón hacer BJHS
Button(
    ventana,
    text="hacer BJHS",
    command=lambda: hacer_BJHS(entry_excel.get(), label_estado)
).grid(row=7, column=1, pady=10)

# Botón hacer tests
Button(
    ventana,
    text="hacer tests",
    command=lambda: hacer_tests(entry_excel.get(), label_estado)
).grid(row=7, column=2, pady=10)

Button(
    ventana,
    text="Listar imágenes",
    command=mostrar_imagenes_en_carrusel
).grid(row=8, column=0, columnspan=3, pady=10)

Button(
    ventana,
    text="Visualizar poro fractal",

    command=lambda: visualizar_poro_fractal(entry_excel.get(), label_estado)
).grid(row=8, column=0, pady=10)

def cargar_configuracion():
    config = configparser.ConfigParser()
    config.read(config_file)
    if "Rutas" in config:
        entry_excel.insert(0, config["Rutas"].get("ruta_excel", ""))


def guardar_configuracion():
    config = configparser.ConfigParser()
    config["Rutas"] = {
        "ruta_excel": entry_excel.get(),
   
    }
    with open(config_file, "w") as configfile:
        config.write(configfile)

ventana.protocol("WM_DELETE_WINDOW", lambda: [guardar_configuracion(), ventana.destroy()])

cargar_configuracion()

ventana.mainloop()
