#Graficos a partir de los dataframes

import openpyxl
import os
import csv
import traceback
from datetime import datetime
from pywinauto import Application, findwindows
import time
import threading
from novawinmng import manejar_novawin, leer_csv_y_crear_dataframe
from pywinauto.keyboard import send_keys
from openpyxl import Workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from pandas import ExcelWriter
from openpyxl import load_workbook
from pandas import ExcelWriter
import subprocess
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.utils.dataframe import dataframe_to_rows

def agregar_dataframe_a_nueva_hoja(archivo_excel, dataframe, nombre_hoja):
    # Cargar el archivo Excel existente
    book = load_workbook(archivo_excel)
    
    # Usar el modo 'append' para evitar sobrescribir el archivo
    with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        # Pasar el objeto book al escritor
        writer._book = book  # Nota: Usamos '_book' en lugar de 'book'
        
        # Escribir el DataFrame en la nueva hoja
        dataframe.to_excel(writer, sheet_name=nombre_hoja, index=False)

def BET_BI(df,Rango_de_Absorpcion, Rango_de_Desorpcion):

    # Verificar si las columnas necesarias existen en el DataFrame
    if 'Volume @ STP' not in df.columns or '1 / [ W((P/Po) - 1) ]' not in df.columns:
        print("Las columnas necesarias no están en el DataFrame.")
        return
    try:
        # Convertir Rango_de_Desorpcion a entero y filtrar
        num_filas = int(Rango_de_Desorpcion)
        ultimo_rango_df = df.tail(num_filas)
        print("Últimos elementos filtrados según Rango de Desorción:")
        print(ultimo_rango_df)
    except ValueError:
        print("El valor de Rango_de_Desorpcion no es válido para realizar el filtrado.")
    
    # Filtrar los rangos y mostrar los datos
    filtered_df_a = ultimo_rango_df[
        (ultimo_rango_df['Relative Pressure'] >= 0.45) & (ultimo_rango_df['Relative Pressure'] <= 0.55)
    ]
    print("Filtrado en rango 0.45-0.55:")
    print(filtered_df_a)

    # Calcular el promedio de la columna 'Volume @ STP'
    if not ultimo_rango_df['Volume @ STP'].isnull().all():  # Verificar si la columna no está vacía
        promedio_volume_stp_a = filtered_df_a['Volume @ STP'].mean()
        print(f"Promedio de 'Volume @ STP': {promedio_volume_stp_a}")
    else:
        print("La columna 'Volume @ STP' está vacía o contiene solo valores nulos.")

    filtered_df_b = ultimo_rango_df[
        (ultimo_rango_df['Relative Pressure'] >= 0.6) & (ultimo_rango_df['Relative Pressure'] <= 0.75)
    ]
    print("Filtrado en rango 0.6-0.75:")
    print(filtered_df_b)

    # Calcular el promedio de la columna 'Volume @ STP'
    if not filtered_df_b['Volume @ STP'].isnull().all():  # Verificar si la columna no está vacía
        promedio_volume_stp_b = filtered_df_b['Volume @ STP'].mean()
        print(f"Promedio de 'Volume @ STP': {promedio_volume_stp_b}")
    else:
        print("La columna 'Volume @ STP' está vacía o contiene solo valores nulos.")

    # Dividir los promedios
    div = promedio_volume_stp_a / promedio_volume_stp_b
    print("La división es: " + str(round(div, 2)))  # Convierte el número a cadena y lo redondea
    if (div >= 1):
        return True
    else:
        return False
    # Filtrar los últimos N elementos según el valor de Rango_de_Desorpcion
   
        
def BET_P(df, Rango_de_Absorpcion, Rango_de_Desorpcion):

    # Verificar si las columnas necesarias existen en el DataFrame
    if 'Volume @ STP' not in df.columns or '1 / [ W((P/Po) - 1) ]' not in df.columns:
        print("Las columnas necesarias no están en el DataFrame.")
        return
    try:
        # Convertir Rango_de_Desorpcion a entero y filtrar
        num_filas = int(Rango_de_Desorpcion)
        ultimo_rango_df = df.tail(num_filas)
        print("Últimos elementos filtrados según Rango de Desorción:")
        print(ultimo_rango_df)
    except ValueError:
        print("El valor de Rango_de_Desorpcion no es válido para realizar el filtrado.")

    # Verificar si las columnas necesarias existen en el DataFrame
    if 'Volume @ STP' not in df.columns or '1 / [ W((P/Po) - 1) ]' not in df.columns:
        print("Las columnas necesarias no están en el DataFrame.")
        return
    
     # Filtrar los rangos y mostrar los datos
    filtered_df_a = ultimo_rango_df[
        (ultimo_rango_df['Relative Pressure'] >= 0.6) & (ultimo_rango_df['Relative Pressure'] <= 0.75)
    ]
    print("Filtrado en rango 0.6-0.75:")
    print(filtered_df_a)

    # Calcular el promedio de la columna 'Volume @ STP'
    if not ultimo_rango_df['Volume @ STP'].isnull().all():  # Verificar si la columna no está vacía
        promedio_volume_stp_a = filtered_df_a['Volume @ STP'].mean()
        print(f"Promedio de 'Volume @ STP': {promedio_volume_stp_a}")
    else:
        print("La columna 'Volume @ STP' está vacía o contiene solo valores nulos.")
    
    try:
        # Convertir Rango_de_Absorpcion a entero y filtrar
        num_filas = int(Rango_de_Absorpcion)
        
        # Obtener los primeros "num_filas" elementos
        primeros_rango_df = df.head(num_filas)
        print("Primeros elementos filtrados según Rango de Rango_de_Absorpcion:")
        print(primeros_rango_df)
    except ValueError:
        print("El valor de Rango_de_Absorpcion no es válido para realizar el filtrado.")
        return
   
    filtered_df_b = ultimo_rango_df[
        (ultimo_rango_df['Relative Pressure'] >= 0.6) & (ultimo_rango_df['Relative Pressure'] <= 0.75)
    ]
    print("Filtrado en rango 0.6-0.75:")
    print(filtered_df_b)

    # Calcular el promedio de la columna 'Volume @ STP'
    if not filtered_df_b['Volume @ STP'].isnull().all():  # Verificar si la columna no está vacía
        promedio_volume_stp_b = filtered_df_b['Volume @ STP'].mean()
        print(f"Promedio de 'Volume @ STP': {promedio_volume_stp_b}")
    else:
        print("La columna 'Volume @ STP' está vacía o contiene solo valores nulos.")

    # Dividir los promedios
    div = promedio_volume_stp_a / promedio_volume_stp_b
    print("La división es: " + str(round(div, 2)))  # Convierte el número a cadena y lo redondea
    if (div >= 1):
        return True
    else:
        return False
    # Filtrar los últimos N elementos según el valor de Rango_de_Desorpcion
def BET_C(df, Rango_de_Absorpcion, Rango_de_Desorpcion):


    # Verificar si las columnas necesarias existen en el DataFrame
    if 'Volume @ STP' not in df.columns or '1 / [ W((P/Po) - 1) ]' not in df.columns:
        print("Las columnas necesarias no están en el DataFrame.")
        return
    try:
        # Convertir Rango_de_Desorpcion a entero y filtrar
        num_filas = int(Rango_de_Desorpcion)
        ultimo_rango_df = df.tail(num_filas)
        print("Últimos elementos filtrados según Rango de Desorción:")
        print(ultimo_rango_df)
    except ValueError:
        print("El valor de Rango_de_Desorpcion no es válido para realizar el filtrado.")

    # Verificar si las columnas necesarias existen en el DataFrame
    if 'Volume @ STP' not in df.columns or '1 / [ W((P/Po) - 1) ]' not in df.columns:
        print("Las columnas necesarias no están en el DataFrame.")
        return
    
     # Filtrar los rangos y mostrar los datos
    filtered_df_a = ultimo_rango_df[
        (ultimo_rango_df['Relative Pressure'] >= 0.95) & (ultimo_rango_df['Relative Pressure'] <= 1.0)
    ]
    print("Filtrado en rango 0.95-1.0:")
    print(filtered_df_a)
   
    try:
        # Convertir Rango_de_Absorpcion a entero y filtrar
        num_filas = int(Rango_de_Absorpcion)
        
        # Obtener los primeros "num_filas" elementos
        primeros_rango_df = df.head(num_filas)
        print("Primeros elementos filtrados según Rango de Rango_de_Absorpcion:")
        print(primeros_rango_df)
    except ValueError:
        print("El valor de Rango_de_Absorpcion no es válido para realizar el filtrado.")
        return
   
    filtered_df_b = ultimo_rango_df[
        (ultimo_rango_df['Relative Pressure'] >= 0.95) & (ultimo_rango_df['Relative Pressure'] <= 1.0)
    ]
    print("Filtrado en rango 0.95-1.0:")
    print(filtered_df_b)

    # Calcular el promedio de la columna 'Volume @ STP'
    if not filtered_df_b['Volume @ STP'].isnull().all():  # Verificar si la columna no está vacía
        promedio_volume_stp_ab = filtered_df_b['Volume @ STP'].mean()
        promedio_volume_stp_ab = promedio_volume_stp_ab + filtered_df_a['Volume @ STP'].mean()
        print(f"Promedio de 'Volume @ STP de valores de Absorbcion y Desorbcion es ': {promedio_volume_stp_ab}")
    else:
        print("La columna 'Volume @ STP' está vacía o contiene solo valores nulos.")

    if (promedio_volume_stp_ab >= 1):
        return True
    else:
        return False
    # Filtrar los últimos N elementos según el valor de Rango_de_Desorpcion
import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import subprocess

def tests_main(df_bet, directorio_guardado):
    import os
    import pandas as pd
    import matplotlib.pyplot as plt

    print("Inicio de tests_main")

    resultados = pd.DataFrame(columns=["Test", "Resultado", "Promedio_A", "Promedio_B", "División"])
    df = df_bet.copy()

    # Calcular diferencia de presión relativa
    df['delta_pressure'] = df['Relative Pressure'].diff()

    Rango_de_Absorpcion = df[df['delta_pressure'] >= 0]
    Rango_de_Desorpcion = df[df['delta_pressure'] < 0]
    num_absorcion = len(Rango_de_Absorpcion)
    num_desorcion = len(Rango_de_Desorpcion)

    # Tests
    resultado_bi = BET_BI(df, num_absorcion, num_desorcion)
    resultado_p = BET_P(df, num_absorcion, num_desorcion)
    resultado_c = BET_C(df, num_absorcion, num_desorcion)

    resultados.loc[len(resultados)] = [
        "BET_BI", "Hay poros cuello de botella" if resultado_bi else "No hay poros cuello de botella", "-", "-", "-"
    ]
    resultados.loc[len(resultados)] = [
        "BET_P", "Hay poros planos" if resultado_p else "No hay poros planos", "-", "-", "-"
    ]
    resultados.loc[len(resultados)] = [
        "BET_C", "Hay poros cilíndricos" if resultado_c else "No hay poros cilíndricos", "-", "-", "-"
    ]

    # Generar gráfico
    plt.hist(df['Volume @ STP'], bins=20, color='blue', alpha=0.7)
    plt.title("Histograma de Volume @ STP")
    plt.xlabel("Volume @ STP")
    plt.ylabel("Frecuencia")
    plt.tight_layout()

    # Guardar imagen temporal
    grafico_path = os.path.join(directorio_guardado, "histograma.png")
    plt.savefig(grafico_path, dpi=300)
    plt.close()

    print("Resultados y gráfico generados.")

    return resultados, grafico_path